from openpyxl import load_workbook

def load_map_with_color(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    grid = []
    for i in range(1, 16):  # 1~15行
        row_data = []
        for j in range(1, 16):  # 1~15列
            cell = ws.cell(row=i, column=j)

            # 读取单元格文字
            val = str(cell.value).strip().upper() if cell.value is not None else ''

            # 读取单元格背景色
            fill_color = cell.fill.fgColor.rgb  # 返回类似 'FF00FF00'
            has_fill = fill_color not in [None, '00000000', 'FFFFFFFF']

            if val == 'S':               # 起点
                row_data.append(2)
            elif val in ['E', '1']:      # 文字标记事件
                row_data.append(1)
            elif has_fill:               # 有填色也算事件
                row_data.append(1)
            else:
                row_data.append(0)       # 空地
        grid.append(row_data)

    return grid


# 测试用
if __name__ == "__main__":
    grid = load_map_with_color("map/ch1.xlsx")
    for row in grid:
        print(row)
